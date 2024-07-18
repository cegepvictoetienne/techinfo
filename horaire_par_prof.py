#-------------------------------------------
# Générateur d'horaire par professeur
#-------------------------------------------
# Ce script génère un fichier markdown à partir d'un fichier Excel contenant les horaires des professeurs.
# Voir le fichier horaire.xlsx.
#-------------------------------------------
# Auteur: Etienne Rivard
# Modifications: Mathieu Fréchette
#-------------------------------------------

from openpyxl import load_workbook

PAR_JOUR = 9
LIGNE_DEBUT_COURS = 3
TITRE_PAGE = "# Horaire des professeurs pour l'automne 2024"
REUNION_DEPARTEMENT = {
    "titre":"Réunion&nbsp;départementale",
    "heure_debut":"8:15",
    "heure_fin":"10:05",
    "span": 2,
    "jour": "Mercredi",
    "local": "C208",
    "prof": "Departement",
    "initiales": "DI"
}

def generer_details_cours(cour, ligne):
    colonne = int(cour["heure_debut"].split(":")[0]) - 7;
    if not cour["groupe"] == "None":
        groupe = f'<span class="cal-class-group">gr. {cour["groupe"]}</span>'
    else:
        groupe = ""

    return f'''<div class="cal-class {cour['prof']}" style="grid-column: {colonne} / span {cour['span']}; grid-row: {ligne};">
                <div class="cal-info">
                    <div class="cal-class-lign-haut"><span class="cal-class-time">{cour["heure_debut"]} - {cour["heure_fin"]}</span>{groupe}</div>
                    <span class="cal-class-title">{cour["titre"]}</span>
                    <div class="cal-class-lign"><span class="cal-class-professor">{cour["prof"]}</span><span class="cal-class-location">{cour["local"]}</span></div>
                </div>
            </div>'''

def generer_horaire_de_excel(chemin_fichier_excel, chemin_fichier_sortie):

    # Lire le contenu du chiffrier Excel et le mettre dans une structure de données
    wb = load_workbook(chemin_fichier_excel)
    ws = wb.active
    
    profs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        jour, heure_debut, heure_fin, span, prof, titre, local, initiales, groupe = row
        if prof not in profs:
            profs[prof] = { "nom": prof, "initiales": initiales, "cours": {} }
        if jour not in profs[prof]["cours"]:
            profs[prof]["cours"][jour] = []
        profs[prof]["cours"][jour].append({
            "heure_debut": heure_debut,
            "heure_fin": heure_fin,
            "span": span,
            "prof": prof,
            "titre": titre,
            "local": local,
            "initiales": initiales,
            "groupe": groupe
        })

    # Début de la génération du contenu Markdown
    
    # Entête des colonnes
    # Colonnne des professeurs
    colonne_prof = '''<div class="cal-column professeur">
    <div></div>
    <div>Professeur</div>'''
    
    # Colonnne des jours
    colonne_lundi = '''<div class="cal-column cal-day lundi"><div class="cal-day-heading">Lundi</div>'''
    colonne_mardi = '''<div class="cal-column cal-day mardi"><div class="cal-day-heading">Mardi</div>'''
    colonne_mercredi = '''<div class="cal-column cal-day mercredi"><div class="cal-day-heading">Mercredi</div>'''
    colonne_jeudi = '''<div class="cal-column cal-day jeudi"><div class="cal-day-heading">Jeudi</div>'''
    colonne_vendredi = '''<div class="cal-column cal-day vendredi"><div class="cal-day-heading">Vendredi</div>'''

    # Ajout des heures à chaque jour
    for heure in range(8, 17):
        colonne_lundi += f'<div class="cal-hour-heading">{heure}</div>'
        colonne_mardi += f'<div class="cal-hour-heading">{heure}</div>'
        colonne_mercredi += f'<div class="cal-hour-heading">{heure}</div>'
        colonne_jeudi += f'<div class="cal-hour-heading">{heure}</div>'
        colonne_vendredi += f'<div class="cal-hour-heading">{heure}</div>'

    # Ajout de la réunion départementale
    colonne_mercredi += f'''<div class="cal-class Departement" style="grid-column: 1 / span 2; grid-row: 3 / span 11;">
                <div class="title"><div>{REUNION_DEPARTEMENT["titre"]}</div></div>
                <div class="cal-info">
                    <span class="cal-class-time">{REUNION_DEPARTEMENT["heure_debut"]} - {REUNION_DEPARTEMENT["heure_fin"]}</span>
                    <span class="cal-class-title">{REUNION_DEPARTEMENT["titre"]}</span>
                    <span class="cal-class-location">{REUNION_DEPARTEMENT["local"]}</span>
                </div>
            </div>'''

    # Ajout des cours à chaque jour
    profs_tries = sorted(profs)
    for prof in profs_tries:
        ligne_prof = profs_tries.index(prof) + LIGNE_DEBUT_COURS
        colonne_prof += f'<div>{profs[prof]["nom"]}</div>'
        for jour in profs[prof]["cours"]:
            for cour in profs[prof]["cours"][jour]:
                match jour:
                    case "Lundi":
                        colonne_lundi += generer_details_cours(cour, ligne_prof)
                    case "Mardi":
                        colonne_mardi += generer_details_cours(cour, ligne_prof)
                    case "Mercredi":
                        colonne_mercredi += generer_details_cours(cour, ligne_prof)
                    case "Jeudi":
                        colonne_jeudi += generer_details_cours(cour, ligne_prof)
                    case "Vendredi":
                        colonne_vendredi += generer_details_cours(cour, ligne_prof)


    # Fermeture des colonnes
    colonne_prof += '</div>'
    colonne_lundi += '</div>'
    colonne_mardi += '</div>'
    colonne_mercredi += '</div>'
    colonne_jeudi += '</div>'
    colonne_vendredi += '</div>'



    contenu_page = f'''
{ TITRE_PAGE }

<div id="cal-calendar">
{ colonne_prof }
{ colonne_lundi }
{ colonne_mardi }
{ colonne_mercredi }
{ colonne_jeudi }
{ colonne_vendredi }
</div>
'''
    with open(chemin_fichier_sortie, 'w', encoding='utf-8') as f:
        f.write(contenu_page)

generer_horaire_de_excel('./template/horaire.xlsx', './wiki/horaire.md')