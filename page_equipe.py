#-------------------------------------------
# Générateur d'horaire par professeur
#-------------------------------------------
# Ce script génère un fichier markdown à partir d'un fichier Excel contenant les informations des professeurs.
# Voir le fichier ./template/equipe.xlsx.
#-------------------------------------------
# Auteur: Etienne Rivard
#-------------------------------------------

from openpyxl import load_workbook

TITRE_PAGE = "# Notre équipe"
ICONE_TEAMS = '''<svg xmlns="http://www.w3.org/2000/svg" width="28" height="28" viewBox="0 0 24 24"><path fill="currentColor" d="M19.19 8.77q-.46 0-.86-.17q-.39-.17-.69-.47t-.47-.69q-.17-.4-.17-.86q0-.45.17-.85t.47-.69q.3-.3.69-.47q.4-.18.86-.17q.45-.01.85.17q.4.17.7.47q.29.29.47.69q.17.4.17.85q0 .46-.17.86q-.17.39-.47.69t-.7.47t-.85.17m0-3.12q-.39 0-.69.27q-.25.27-.25.66t.25.67q.3.25.69.25t.66-.25q.28-.25.28-.67q0-.39-.28-.66q-.27-.27-.66-.27M22 10.33V15q0 .63-.24 1.2q-.26.57-.67.99q-.43.43-1 .67q-.59.25-1.21.25q-.38 0-.76-.11q-.39-.07-.71-.25q-.24.79-.71 1.44t-1.1 1.11t-1.39.7q-.76.27-1.58.27q-.96 0-1.81-.33q-.82-.33-1.5-.94q-.66-.57-1.09-1.36q-.44-.8-.57-1.74H2.83q-.33 0-.59-.25q-.24-.24-.24-.58V7.73q0-.34.24-.59q.26-.24.59-.24H10q-.29-.6-.29-1.25q0-.61.23-1.15q.22-.5.62-.92q.4-.39.94-.62q.5-.23 1.12-.23q.61 0 1.14.23t.93.62q.4.42.62.92q.23.54.23 1.15q0 .6-.23 1.14q-.22.53-.62.92q-.4.4-.93.63t-1.14.23q-.15 0-.31-.02q-.15-.02-.31-.05v.9h9.06q.39 0 .67.27q.27.27.27.66M12.63 4q-.35 0-.63.11q-.33.13-.56.36q-.22.23-.35.53q-.13.31-.13.65q0 .35.13.65t.35.53q.23.22.56.36q.28.13.63.13q.34 0 .64-.13q.3-.14.53-.36q.23-.23.36-.53q.14-.3.14-.65q0-.34-.14-.65q-.13-.3-.36-.53t-.53-.36q-.3-.11-.64-.11m-4.85 6.18h1.88V8.62H4.34v1.56h1.88v5h1.56m8.6 1.09v-5.62H12v5.42q0 .34-.24.58q-.26.25-.59.25H8.92q.13.67.47 1.25q.34.57.82.99q.48.41 1.1.65q.61.21 1.32.21q.77 0 1.45-.27q.68-.3 1.2-.81q.51-.51.8-1.19q.3-.68.3-1.46M20.75 15v-4.35h-3.12v5.71q.25.25.57.38q.3.12.68.12q.39 0 .73-.15t.59-.4q.26-.25.4-.6q.15-.34.15-.71"/></svg>'''

def generer_card(prof):
    nom, poste, avatar, courriel, telephone, bureau, teams = prof
    return f'''
<div class="team-card">

    <div class="team-card-header"></div>
    <div class="team-card-image">
        <img src="{avatar}" alt="{nom}">
    </div>
        
    <div class="team-card-info">
        <h2 class="team-card-name">{nom}</h2>
        <div class="team-card-poste">{poste if poste is not None else ''}</div>
        <div class="team-card-contact">
            <div><i class="fa-solid fa-envelope"></i><span>{courriel}</span></div>
            <div><i class="fa-solid fa-phone"></i><span>{telephone}</span></div>
            <div><i class="fa-solid fa-building"></i><span>{bureau}</span></div>
            <div class="team-contact-teams">
                <a href="{teams}">
                    {ICONE_TEAMS}
                </a>
            </div>
        </div>
    </div>

</div>
    '''

def generer_page_equipe(chemin_fichier_excel, chemin_fichier_sortie):

    # Lire le contenu du chiffrier Excel
    wb = load_workbook(chemin_fichier_excel)
    ws = wb.active
    
    contenu_page = f'''
{ TITRE_PAGE }
<div class="team-section">
    '''
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        contenu_page += generer_card(row)

    contenu_page += f'''
</div>
    '''
    

    with open(chemin_fichier_sortie, 'w', encoding='utf-8') as f:
        f.write(contenu_page)

generer_page_equipe('./template/equipe.xlsx', './wiki/equipe.md')